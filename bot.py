from extract import extractItems,driverFunc
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import threading
from time import sleep
import datetime



def getItemsFromExcel():
    itemsWithPricesDict = {}
    wb = load_workbook('itemsWithPrices.xlsx')
    ws = wb['Sheet1']
    itemsColumn = ws['A']
    pricesColumn = ws['B']
    for i in range(len(itemsColumn)):
        itemsWithPricesDict[itemsColumn[i].value] = pricesColumn[i].value

    #print(itemsWithPricesDict)    
    return itemsWithPricesDict

def bid(timeRemainingInSeconds,lotNumber):
    
    itemsWithPricesDict = getItemsFromExcel()
    #print(itemsWithPricesDict.get(lotNumber.strip()))
    bidPrice = int(itemsWithPricesDict.get(lotNumber.strip()))
    sleep(timeRemainingInSeconds -140)
    driver = driverFunc()
    driver.get(lotNumber.strip())
    sleep(5)
    #listOfItems = extractItems(driver)
    currentBid = None
    bidButton = None
    # for item in listOfItems:
    #     if(item.get('lotNumber').strip() == lotNumber.strip()):
    #         print(item.get('lotNumber').strip(),lotNumber.strip())
    #         currentBid = item.get('currentBid')
    #         bidButton = item.get('bidNowBtn')
    #         break
    bidButton = driver.find_element(by=By.XPATH, value='//*[@class="btn btn-primary btn-shadow btn-block"]')
    currentBid = float(driver.find_element(by=By.XPATH, value='//*[@class="h1 font-weight-normal text-accent mb-0 mr-4"]/span').text.replace('$',''))
    driver.execute_script("return arguments[0].scrollIntoView();", bidButton)
    sleep(1)
    if(bidPrice >= currentBid):
        bidButton.click()
        try:
            takeRiskBtn = driver.find_element(by=By.XPATH,value=('//*[@class="btn btn-danger mx-1 my-2 btn-block"]'))
            takeRiskBtn.click()
            
        except:
            try:
                registerToBidBtn = driver.find_element(by=By.XPATH,value=('//*[@class="btn btn-block btn-primary"]'))
                registerToBidBtn.click()
                try:
                    takeRiskBtn = driver.find_element(by=By.XPATH,value=('//*[@class="btn btn-danger mx-1 my-2 btn-block"]'))
                    takeRiskBtn.click()
                except Exception as e:
                    print('take risk btn not found after register to bid btn')
                        
            except Exception as e:
                print('regiter to bid btn not found')
    sleep(4)
    driver.close()

def start():
    threads = []
    driver = driverFunc()
    listOfItems = extractItems(driver)
    driver.close()
    #seconds = 180
    for item in listOfItems:
        time = item.get('timeRemaining')
        lotNumber = item.get('lotNumber').strip()
        #print(time)
        #print(time%(24*60*60),(time%(24*60*60))*24 - time%(24*60),(time%(24*60*60))*24*60 - time%60,(time%(24*60*60))*24*60*60 - time)
        #print(lotNumber)
        threads.append(threading.Thread(target=bid,args=(time,lotNumber)))
        #seconds = seconds + 60

    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join()

start()